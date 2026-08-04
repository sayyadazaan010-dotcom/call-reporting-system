import os
import io
import sqlite3
from datetime import datetime, date
from functools import wraps

from flask import Flask, render_template, request, redirect, url_for, session, flash, send_file, g
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.lib.units import mm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet

BASE_DIR = os.path.abspath(os.path.dirname(__file__))
DB_PATH = os.path.join(BASE_DIR, 'callreports.db')
PROFILE_UPLOAD_DIR = os.path.join(BASE_DIR, 'static', 'uploads', 'profiles')
os.makedirs(PROFILE_UPLOAD_DIR, exist_ok=True)
ALLOWED_PHOTO_EXT = {'png', 'jpg', 'jpeg', 'webp'}
PETROL_RATE_PER_KM = 7.0

# Fixed travel routes (label, km). Selecting one auto-fills location + KM + amount.
PETROL_ROUTES = [
    {
        "label": "Porvorim Arena → Bicholim Arena (Route 1)",
        "location": "Porvorim Arena to Bicholim Arena (Route 1)",
        "km": 31.0,
        "maps_url": "https://maps.app.goo.gl/ZFD6ebDq7SZEZdFu5?g_st=ac",
    },
    {
        "label": "Porvorim Arena → Bicholim Arena (Route 2)",
        "location": "Porvorim Arena to Bicholim Arena (Route 2)",
        "km": 38.0,
        "maps_url": "https://maps.app.goo.gl/A8bU3qGDN7Axzvzt8?g_st=ac",
    },
    {
        "label": "Bicholim Arena → Bicholim Nexa Studio",
        "location": "Bicholim Arena to Bicholim Nexa Studio",
        "km": 5.7,
        "maps_url": "https://maps.app.goo.gl/RwrfQTb34GPiRbJF7?g_st=ac",
    },
    {
        "label": "Porvorim Arena → Mapusa",
        "location": "Porvorim Arena to Mapusa",
        "km": 14.0,
        "maps_url": "https://maps.app.goo.gl/B1kyKHvXMC2mkoV87?g_st=ac",
    },
    {
        "label": "Porvorim Arena → Nexa Sales",
        "location": "Porvorim Arena to Nexa Sales",
        "km": 5.6,
        "maps_url": "https://maps.app.goo.gl/jS8pUwUYzruQWzoy7?g_st=ac",
    },
    {
        "label": "Porvorim Arena → Nexa Workshop",
        "location": "Porvorim Arena to Nexa Workshop",
        "km": 3.7,
        "maps_url": "https://maps.app.goo.gl/SVZaQNJauceKEUEx7?g_st=ac",
    },
    {
        "label": "Porvorim Arena → Nexa Workshop + Nexa Sales + Porvorim",
        "location": "Porvorim Arena to Nexa Workshop and Nexa Sales and Porvorim",
        "km": 9.0,
        "maps_url": "https://maps.app.goo.gl/kgLvEntipTF8P7hk8?g_st=ac",
    },
    {
        "label": "Porvorim Arena → Verna Sai Service Arena",
        "location": "Porvorim Arena to Verna Sai Service Arena",
        "km": 27.0,
        "maps_url": "https://maps.app.goo.gl/fPbGmUri9NT92tnT9?g_st=ac",
    },
]

app = Flask(__name__)
app.config['SECRET_KEY'] = os.environ.get('SECRET_KEY', 'change-this-secret-key-in-production')
app.config['MAX_CONTENT_LENGTH'] = 8 * 1024 * 1024  # 8 MB upload cap


@app.template_filter('fmt_dt')
def fmt_dt(value, fmt='%d %b %Y %H:%M'):
    if not value:
        return '-'
    try:
        return datetime.fromisoformat(value).strftime(fmt)
    except (ValueError, TypeError):
        return value


@app.template_filter('fmt_date')
def fmt_date(value, fmt='%d %b %Y'):
    if not value:
        return '-'
    try:
        return datetime.strptime(value, '%Y-%m-%d').strftime(fmt)
    except (ValueError, TypeError):
        return value

LOCATIONS = [
    "Porvorim Arena",
    "Nexa Sales",
    "Nexa Workshop",
    "Pilerne",
    "Panjim",
    "Stockyard",
    "Mapusa",
    "Nexa Thivim",
    "Colvale",
    "Bicholim Arena",
    "Bicholim Studio",
    "Sai service Arena (verna)",
    "Ponda",
    "True value (verna)",
    "Nexa workshop (verna)",
    "Cuncolim",
    "Cuncolim R outlet",
    "Vasco",
]

DEPARTMENTS = [
    "IT",
    "Account",
    "HR",
    "Insurance",
    "Sales",
    "Service",
    "Bodyshop",
    "Billing",
    "Back office",
    "True value",
    "Accessories",
    "MDS",
]
LEAVE_TYPES = ["Sick Leave", "Casual Leave", "Earned Leave", "Emergency Leave", "Other"]

TONER_TYPES = ["12A", "88A", "Other"]
DEFAULT_CCTV_PURPOSE = "To view CCTV footage"
REACHED_DELAY_SECONDS = 24 * 60 * 60  # "Reached" status can only be set 1 day after submission


# ---------------------------------------------------------------------------
# Database helpers
# ---------------------------------------------------------------------------

def get_db():
    if 'db' not in g:
        g.db = sqlite3.connect(DB_PATH)
        g.db.row_factory = sqlite3.Row
        g.db.execute("PRAGMA foreign_keys = ON")
    return g.db


@app.teardown_appcontext
def close_db(exception=None):
    db = g.pop('db', None)
    if db is not None:
        db.close()


def init_db():
    db = sqlite3.connect(DB_PATH)
    db.execute("""
        CREATE TABLE IF NOT EXISTS user (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            username TEXT UNIQUE NOT NULL,
            full_name TEXT NOT NULL,
            email TEXT,
            password_hash TEXT NOT NULL,
            role TEXT NOT NULL DEFAULT 'employee',
            department TEXT,
            is_active INTEGER NOT NULL DEFAULT 1,
            created_at TEXT,
            last_login TEXT
        )
    """)
    db.execute("""
        CREATE TABLE IF NOT EXISTS report (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL,
            report_date TEXT NOT NULL,
            customer_name TEXT,
            location TEXT,
            department TEXT,
            mode TEXT,
            status TEXT,
            issue TEXT,
            entry_type TEXT DEFAULT 'Visit',
            leave_type TEXT,
            created_at TEXT,
            FOREIGN KEY (user_id) REFERENCES user (id)
        )
    """)
    db.execute("""
        CREATE TABLE IF NOT EXISTS petrol_expense (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL,
            expense_date TEXT NOT NULL,
            location TEXT,
            km REAL,
            amount REAL NOT NULL,
            created_at TEXT,
            FOREIGN KEY (user_id) REFERENCES user (id)
        )
    """)
    db.execute("""
        CREATE TABLE IF NOT EXISTS toner_entry (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL,
            entry_date TEXT NOT NULL,
            location TEXT,
            department TEXT,
            handover_to TEXT,
            toner_type TEXT,
            quantity INTEGER DEFAULT 1,
            created_at TEXT,
            FOREIGN KEY (user_id) REFERENCES user (id)
        )
    """)
    db.execute("""
        CREATE TABLE IF NOT EXISTS cctv_entry (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL,
            entry_date TEXT NOT NULL,
            viewer_name TEXT,
            department TEXT,
            location TEXT,
            purpose TEXT,
            remark TEXT,
            created_at TEXT,
            FOREIGN KEY (user_id) REFERENCES user (id)
        )
    """)
    db.execute("""
        CREATE TABLE IF NOT EXISTS toner_requirement (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL,
            entry_date TEXT NOT NULL,
            requester_name TEXT,
            needed_by TEXT,
            department TEXT,
            location TEXT,
            toner_type TEXT,
            quantity INTEGER DEFAULT 1,
            remark TEXT,
            status TEXT DEFAULT 'open',
            created_at TEXT,
            FOREIGN KEY (user_id) REFERENCES user (id)
        )
    """)
    db.execute("""
        CREATE TABLE IF NOT EXISTS visitor_entry (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL,
            entry_date TEXT NOT NULL,
            location TEXT,
            visitor_name TEXT,
            purpose TEXT,
            phone TEXT,
            created_at TEXT,
            FOREIGN KEY (user_id) REFERENCES user (id)
        )
    """)
    db.commit()

    # Safe migration: add photo_filename column if this DB predates it
    existing_cols = [row[1] for row in db.execute("PRAGMA table_info(user)").fetchall()]
    if 'photo_filename' not in existing_cols:
        db.execute("ALTER TABLE user ADD COLUMN photo_filename TEXT")
        db.commit()

    # Safe migration: add customer_name to report if this DB predates it
    report_cols = [row[1] for row in db.execute("PRAGMA table_info(report)").fetchall()]
    if 'customer_name' not in report_cols:
        db.execute("ALTER TABLE report ADD COLUMN customer_name TEXT")
        db.commit()

    if 'reached' not in report_cols:
        db.execute("ALTER TABLE report ADD COLUMN reached TEXT")
        db.commit()

    if 'entry_type' not in report_cols:
        db.execute("ALTER TABLE report ADD COLUMN entry_type TEXT DEFAULT 'Visit'")
        db.commit()

    if 'leave_type' not in report_cols:
        db.execute("ALTER TABLE report ADD COLUMN leave_type TEXT")
        db.commit()

    # Safe migrations on petrol_expense
    petrol_cols = [row[1] for row in db.execute("PRAGMA table_info(petrol_expense)").fetchall()]
    if 'reported_at' not in petrol_cols:
        db.execute("ALTER TABLE petrol_expense ADD COLUMN reported_at TEXT")
        db.commit()
    if 'day_type' not in petrol_cols:
        db.execute("ALTER TABLE petrol_expense ADD COLUMN day_type TEXT DEFAULT 'Working'")
        db.commit()

    # Safe migrations on toner_entry
    toner_cols = [row[1] for row in db.execute("PRAGMA table_info(toner_entry)").fetchall()]
    if 'reached' not in toner_cols:
        db.execute("ALTER TABLE toner_entry ADD COLUMN reached TEXT")
        db.commit()
    if 'quantity' not in toner_cols:
        db.execute("ALTER TABLE toner_entry ADD COLUMN quantity INTEGER DEFAULT 1")
        db.commit()

    # Safe migrations on toner_requirement
    req_cols = [row[1] for row in db.execute("PRAGMA table_info(toner_requirement)").fetchall()]
    for col, ddl in [
        ('needed_by', "ALTER TABLE toner_requirement ADD COLUMN needed_by TEXT"),
        ('department', "ALTER TABLE toner_requirement ADD COLUMN department TEXT"),
        ('quantity', "ALTER TABLE toner_requirement ADD COLUMN quantity INTEGER DEFAULT 1"),
        ('status', "ALTER TABLE toner_requirement ADD COLUMN status TEXT DEFAULT 'open'"),
    ]:
        if col not in req_cols:
            db.execute(ddl)
            db.commit()

    existing = db.execute("SELECT id FROM user WHERE username = ?", ('admin',)).fetchone()
    if not existing:
        db.execute(
            "INSERT INTO user (username, full_name, email, password_hash, role, department, is_active, created_at) "
            "VALUES (?, ?, ?, ?, ?, ?, 1, ?)",
            ('admin', 'Azaan', '', generate_password_hash('admin'), 'admin', 'IT Department',
             datetime.utcnow().isoformat())
        )
        db.commit()
        print("Default admin created -> username: admin / password: admin (please change it after first login)")
    db.close()


# ---------------------------------------------------------------------------
# Auth helpers
# ---------------------------------------------------------------------------

def login_required(f):
    @wraps(f)
    def wrapped(*args, **kwargs):
        if 'user_id' not in session:
            return redirect(url_for('login'))
        # Restricted toner-requirement users may only use their portal
        if session.get('role') == 'toner_req' and request.endpoint not in (
            'toner_req_portal', 'logout', 'static'
        ):
            return redirect(url_for('toner_req_portal'))
        return f(*args, **kwargs)
    return wrapped


def admin_required(f):
    @wraps(f)
    def wrapped(*args, **kwargs):
        if session.get('role') != 'admin':
            flash('Administrator access required.', 'error')
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return wrapped


def current_user():
    uid = session.get('user_id')
    if not uid:
        return None
    db = get_db()
    return db.execute("SELECT * FROM user WHERE id = ?", (uid,)).fetchone()


def parse_date(s, default=None):
    try:
        return datetime.strptime(s, '%Y-%m-%d').date()
    except (ValueError, TypeError):
        return default if default is not None else date.today()


def reached_eligible(created_at):
    """The 'Reached' toggle only appears once a day has passed since submission,
    so freshly submitted reports don't prompt for it immediately."""
    if not created_at:
        return True
    try:
        created_dt = datetime.fromisoformat(created_at)
    except (ValueError, TypeError):
        return True
    return (datetime.utcnow() - created_dt).total_seconds() >= REACHED_DELAY_SECONDS


# ---------------------------------------------------------------------------
# Routes: Auth
# ---------------------------------------------------------------------------

def home_for_role(role):
    """Redirect target after login based on role."""
    if role == 'admin':
        return url_for('admin_dashboard')
    if role == 'toner_req':
        return url_for('toner_req_portal')
    return url_for('new_report')


@app.route('/', methods=['GET'])
def index():
    if 'user_id' in session:
        return redirect(home_for_role(session.get('role')))
    return redirect(url_for('login'))


@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '')
        db = get_db()
        user = db.execute("SELECT * FROM user WHERE username = ?", (username,)).fetchone()
        if user and user['is_active'] and check_password_hash(user['password_hash'], password):
            session.clear()
            session['user_id'] = user['id']
            session['role'] = user['role']
            session['full_name'] = user['full_name']
            db.execute("UPDATE user SET last_login = ? WHERE id = ?", (datetime.utcnow().isoformat(), user['id']))
            db.commit()
            return redirect(home_for_role(user['role']))
        flash('Invalid username or password.', 'error')
    return render_template('login.html')


@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))


@app.route('/change-password', methods=['GET', 'POST'])
@login_required
def change_password():
    user = current_user()
    if request.method == 'POST':
        current_pw = request.form.get('current_password', '')
        new_pw = request.form.get('new_password', '')
        confirm_pw = request.form.get('confirm_password', '')
        if not check_password_hash(user['password_hash'], current_pw):
            flash('Current password is incorrect.', 'error')
        elif len(new_pw) < 4:
            flash('New password must be at least 4 characters.', 'error')
        elif new_pw != confirm_pw:
            flash('New passwords do not match.', 'error')
        else:
            db = get_db()
            db.execute("UPDATE user SET password_hash = ? WHERE id = ?",
                       (generate_password_hash(new_pw), user['id']))
            db.commit()
            flash('Password updated successfully.', 'success')
            return redirect(url_for('admin_dashboard') if user['role'] == 'admin' else url_for('new_report'))
    return render_template('change_password.html', user=user, active_page='change_password')


@app.route('/settings/photo', methods=['POST'])
@login_required
def upload_photo():
    user = current_user()
    file = request.files.get('photo')
    if not file or file.filename == '':
        flash('Please choose an image to upload.', 'error')
        return redirect(url_for('change_password'))

    ext = file.filename.rsplit('.', 1)[-1].lower() if '.' in file.filename else ''
    if ext not in ALLOWED_PHOTO_EXT:
        flash('Only PNG, JPG, JPEG or WEBP images are allowed.', 'error')
        return redirect(url_for('change_password'))

    filename = secure_filename(f"user_{user['id']}.{ext}")
    file.save(os.path.join(PROFILE_UPLOAD_DIR, filename))

    db = get_db()
    db.execute("UPDATE user SET photo_filename = ? WHERE id = ?",
               (f"uploads/profiles/{filename}", user['id']))
    db.commit()
    flash('Profile photo updated.', 'success')
    return redirect(url_for('change_password'))


# ---------------------------------------------------------------------------
# Routes: Employee
# ---------------------------------------------------------------------------

@app.route('/new-report', methods=['GET', 'POST'])
@login_required
def new_report():
    user = current_user()
    db = get_db()
    if request.method == 'POST':
        report_date = parse_date(request.form.get('report_date'))
        entry_type = request.form.get('entry_type', 'Visit')
        leave_type = request.form.get('leave_type', '').strip() if entry_type == 'Leave' else None

        if entry_type in ('Holiday', 'Leave'):
            # Holiday / Leave entries don't need customer/location/department details
            db.execute(
                "INSERT INTO report (user_id, report_date, customer_name, location, department, mode, status, issue, entry_type, leave_type, created_at) "
                "VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
                (user['id'], report_date.isoformat(), '', '', '', '', '', '',
                 entry_type, leave_type, datetime.utcnow().isoformat())
            )
        else:
            db.execute(
                "INSERT INTO report (user_id, report_date, customer_name, location, department, mode, status, issue, entry_type, leave_type, created_at) "
                "VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
                (user['id'], report_date.isoformat(), request.form.get('customer_name', '').strip(),
                 request.form.get('location'), request.form.get('department'),
                 request.form.get('mode'), request.form.get('status'), request.form.get('issue'),
                 'Visit', None, datetime.utcnow().isoformat())
            )
        db.commit()
        flash('Report submitted successfully.', 'success')
        # Keep the same date selected for the next entry instead of resetting to today
        return redirect(url_for('new_report', date=report_date.isoformat()))

    total_reports = db.execute("SELECT COUNT(*) c FROM report WHERE user_id = ?", (user['id'],)).fetchone()['c']
    pending_reports = db.execute(
        "SELECT COUNT(*) c FROM report WHERE user_id = ? AND status = 'Pending'", (user['id'],)
    ).fetchone()['c']

    # A fresh page load / refresh with no ?date= always shows today's date
    selected_date = parse_date(request.args.get('date')) if request.args.get('date') else date.today()

    return render_template(
        'employee_dashboard.html',
        user=user,
        locations=LOCATIONS,
        departments=DEPARTMENTS,
        leave_types=LEAVE_TYPES,
        today=selected_date.isoformat(),
        total_reports=total_reports,
        pending_reports=pending_reports,
        active_page='new_report',
    )


@app.route('/my-reports')
@login_required
def my_reports():
    user = current_user()
    db = get_db()
    reports = db.execute(
        "SELECT * FROM report WHERE user_id = ? ORDER BY report_date DESC", (user['id'],)
    ).fetchall()
    return render_template(
        'my_reports.html',
        user=user,
        reports=reports,
        locations=LOCATIONS,
        departments=DEPARTMENTS,
        active_page='my_reports',
    )


@app.route('/reports/<int:report_id>/edit', methods=['POST'])
@login_required
def edit_own_report(report_id):
    """Employee can edit their own call report; admin can edit any."""
    user = current_user()
    db = get_db()
    r = db.execute("SELECT * FROM report WHERE id = ?", (report_id,)).fetchone()
    if not r or (r['user_id'] != user['id'] and user['role'] != 'admin'):
        flash('Report not found.', 'error')
        return redirect(url_for('admin_reports') if user['role'] == 'admin' else url_for('my_reports'))

    report_date = parse_date(request.form.get('report_date'))
    db.execute(
        "UPDATE report SET report_date=?, customer_name=?, location=?, department=?, mode=?, status=?, issue=? "
        "WHERE id=?",
        (report_date.isoformat(), request.form.get('customer_name', '').strip(),
         request.form.get('location', '').strip(), request.form.get('department', '').strip(),
         request.form.get('mode'), request.form.get('status'), request.form.get('issue'), report_id)
    )
    db.commit()
    flash('Report updated successfully.', 'success')
    if user['role'] == 'admin':
        return redirect(url_for('admin_reports'))
    return redirect(url_for('my_reports'))


@app.route('/reports/<int:report_id>/delete', methods=['POST'])
@login_required
def delete_own_report(report_id):
    """Employee can delete own report only; admin can delete any."""
    user = current_user()
    db = get_db()
    r = db.execute("SELECT * FROM report WHERE id = ?", (report_id,)).fetchone()
    if not r or (r['user_id'] != user['id'] and user['role'] != 'admin'):
        flash('Report not found or not allowed.', 'error')
        return redirect(url_for('admin_reports') if user['role'] == 'admin' else url_for('my_reports'))
    db.execute("DELETE FROM report WHERE id = ?", (report_id,))
    db.commit()
    flash('Report deleted.', 'success')
    if user['role'] == 'admin':
        return redirect(url_for('admin_reports'))
    return redirect(url_for('my_reports'))


@app.route('/reports/<int:report_id>/set-reached', methods=['POST'])
@login_required
def set_reached(report_id):
    user = current_user()
    db = get_db()
    r = db.execute("SELECT * FROM report WHERE id = ?", (report_id,)).fetchone()
    next_page = request.form.get('next', 'my_reports')
    redirect_target = url_for('admin_reports') if (next_page == 'admin_reports' and user['role'] == 'admin') \
        else url_for('my_reports')

    if not r or (r['user_id'] != user['id'] and user['role'] != 'admin'):
        flash('Report not found.', 'error')
        return redirect(redirect_target)

    if not reached_eligible(r['created_at']):
        flash('Reached status can only be set a day after the report was submitted.', 'error')
        return redirect(redirect_target)

    value = request.form.get('reached_value')
    if value not in ('Yes', 'No'):
        flash('Invalid selection.', 'error')
        return redirect(redirect_target)

    db.execute("UPDATE report SET reached = ? WHERE id = ?", (value, report_id))
    db.commit()
    flash('Reached status updated.', 'success')
    return redirect(redirect_target)


# ---------------------------------------------------------------------------
# Routes: Petrol Expenses
# ---------------------------------------------------------------------------

@app.route('/petrol-expenses', methods=['GET', 'POST'])
@login_required
def petrol_expenses():
    user = current_user()
    db = get_db()

    if request.method == 'POST':
        expense_date = parse_date(request.form.get('expense_date'))
        reported_at = request.form.get('reported_at', '').strip()
        entry_mode = request.form.get('entry_mode', 'stay').strip()
        location = request.form.get('location', '').strip()
        km_raw = request.form.get('km', '').strip()
        amount_raw = request.form.get('amount', '').strip()

        # Sunday / weekly holiday — mark day only, no travel fields required
        if entry_mode == 'sunday':
            day_type = 'Sunday'
            reported_at = ''
            location = 'SUNDAY — Weekly Holiday / Non-working day'
            km = 0.0
            amount = 0.0
        elif entry_mode == 'travel':
            day_type = 'Working'
            reported_at = ''  # not used in travel mode
            km = None
            try:
                km = float(km_raw) if km_raw else None
            except ValueError:
                km = None
            try:
                amount = float(amount_raw) if amount_raw else (km * PETROL_RATE_PER_KM if km else 0)
            except ValueError:
                amount = km * PETROL_RATE_PER_KM if km else 0
            if not location:
                flash('Please select or type Location Visited.', 'error')
                return redirect(url_for('petrol_expenses'))
        else:
            # Only Reported At (stay)
            day_type = 'Working'
            if not reported_at:
                flash('Please select or type Reported At location.', 'error')
                return redirect(url_for('petrol_expenses'))
            location = location or f"Reported at {reported_at}"
            km = 0.0
            amount = 0.0

        db.execute(
            "INSERT INTO petrol_expense (user_id, expense_date, reported_at, location, km, amount, day_type, created_at) "
            "VALUES (?, ?, ?, ?, ?, ?, ?, ?)",
            (user['id'], expense_date.isoformat(), reported_at, location, km, amount, day_type,
             datetime.utcnow().isoformat())
        )
        db.commit()
        flash('Petrol expense entry added.', 'success')
        return redirect(url_for('petrol_expenses', date=expense_date.isoformat()))

    selected_date = parse_date(request.args.get('date')) if request.args.get('date') else date.today()

    entries = db.execute(
        "SELECT * FROM petrol_expense WHERE user_id = ? ORDER BY expense_date DESC, id DESC", (user['id'],)
    ).fetchall()
    total_amount = sum(e['amount'] or 0 for e in entries)

    return render_template(
        'petrol_expenses.html',
        user=user,
        entries=entries,
        total_amount=total_amount,
        locations=LOCATIONS,
        petrol_routes=PETROL_ROUTES,
        today=selected_date.isoformat(),
        rate_per_km=PETROL_RATE_PER_KM,
        active_page='petrol_expenses',
    )


@app.route('/petrol-expenses/<int:entry_id>/edit', methods=['POST'])
@login_required
def edit_petrol_expense(entry_id):
    """Employee can edit own petrol entry; admin can edit any."""
    user = current_user()
    db = get_db()
    e = db.execute("SELECT * FROM petrol_expense WHERE id = ?", (entry_id,)).fetchone()
    redirect_target = url_for('admin_petrol_expenses') if user['role'] == 'admin' else url_for('petrol_expenses')
    if not e or (e['user_id'] != user['id'] and user['role'] != 'admin'):
        flash('Entry not found.', 'error')
        return redirect(redirect_target)

    expense_date = parse_date(request.form.get('expense_date'))
    day_type = request.form.get('day_type', 'Working').strip()
    if day_type not in ('Working', 'Sunday'):
        day_type = 'Working'
    reported_at = request.form.get('reported_at', '').strip()
    location = request.form.get('location', '').strip()
    km_raw = request.form.get('km', '').strip()
    amount_raw = request.form.get('amount', '').strip()

    if day_type == 'Sunday':
        reported_at = reported_at or ''
        location = location or 'SUNDAY — Weekly Holiday / Non-working day'
        km = 0.0
        amount = 0.0
    else:
        km = None
        try:
            km = float(km_raw) if km_raw else 0.0
        except ValueError:
            km = 0.0
        try:
            amount = float(amount_raw) if amount_raw else (km * PETROL_RATE_PER_KM if km else 0)
        except ValueError:
            amount = km * PETROL_RATE_PER_KM if km else 0
        if not location and reported_at:
            location = f"Reported at {reported_at}"
        if not location:
            flash('Location is required.', 'error')
            return redirect(redirect_target)

    db.execute(
        "UPDATE petrol_expense SET expense_date=?, reported_at=?, location=?, km=?, amount=?, day_type=? WHERE id=?",
        (expense_date.isoformat(), reported_at, location, km, amount, day_type, entry_id)
    )
    db.commit()
    flash('Petrol expense updated successfully.', 'success')
    return redirect(redirect_target)


@app.route('/petrol-expenses/<int:entry_id>/delete', methods=['POST'])
@login_required
def delete_petrol_expense(entry_id):
    user = current_user()
    db = get_db()
    e = db.execute("SELECT * FROM petrol_expense WHERE id = ?", (entry_id,)).fetchone()
    redirect_target = url_for('admin_petrol_expenses') if user['role'] == 'admin' else url_for('petrol_expenses')
    if not e or (e['user_id'] != user['id'] and user['role'] != 'admin'):
        flash('Entry not found.', 'error')
        return redirect(redirect_target)
    db.execute("DELETE FROM petrol_expense WHERE id = ?", (entry_id,))
    db.commit()
    flash('Entry deleted.', 'success')
    return redirect(redirect_target)


@app.route('/admin/petrol-expenses')
@admin_required
def admin_petrol_expenses():
    db = get_db()
    rows = db.execute("""
        SELECT petrol_expense.*, user.full_name AS employee_name
        FROM petrol_expense JOIN user ON petrol_expense.user_id = user.id
        ORDER BY petrol_expense.expense_date DESC, petrol_expense.id DESC
    """).fetchall()
    total_amount = sum(r['amount'] or 0 for r in rows)
    return render_template(
        'admin_petrol_expenses.html',
        user=current_user(),
        entries=rows,
        total_amount=total_amount,
        locations=LOCATIONS,
        rate_per_km=PETROL_RATE_PER_KM,
        active_page='admin_petrol_expenses',
    )


# ---------------------------------------------------------------------------
# Routes: Toner Management
# ---------------------------------------------------------------------------

@app.route('/toner', methods=['GET', 'POST'])
@login_required
def toner_management():
    user = current_user()
    db = get_db()

    if request.method == 'POST':
        entry_date = parse_date(request.form.get('entry_date'))
        toner_type = request.form.get('toner_type', '').strip()
        if toner_type == 'Other':
            toner_type = request.form.get('toner_type_other', '').strip() or 'Other'
        try:
            qty = int(request.form.get('quantity') or 1)
        except ValueError:
            qty = 1
        if qty < 1:
            qty = 1

        db.execute(
            "INSERT INTO toner_entry (user_id, entry_date, location, department, handover_to, toner_type, quantity, created_at) "
            "VALUES (?, ?, ?, ?, ?, ?, ?, ?)",
            (user['id'], entry_date.isoformat(), request.form.get('location', '').strip(),
             request.form.get('department', '').strip(), request.form.get('handover_to', '').strip(),
             toner_type, qty, datetime.utcnow().isoformat())
        )
        db.commit()
        flash('Toner entry recorded.', 'success')
        return redirect(url_for('toner_management', date=entry_date.isoformat()))

    selected_date = parse_date(request.args.get('date')) if request.args.get('date') else date.today()
    rows = db.execute(
        "SELECT * FROM toner_entry WHERE user_id = ? ORDER BY entry_date DESC, id DESC", (user['id'],)
    ).fetchall()
    entries = []
    pending_followups = []
    for r in rows:
        d = dict(r)
        d['reached_eligible'] = reached_eligible(r['created_at'])
        entries.append(d)
        if d['reached_eligible'] and not (r['reached'] or '').strip():
            pending_followups.append(d)

    # Open toner requirements from any location → popup for all staff
    open_reqs = db.execute("""
        SELECT toner_requirement.*, user.full_name AS employee_name
        FROM toner_requirement JOIN user ON toner_requirement.user_id = user.id
        WHERE COALESCE(toner_requirement.status, 'open') = 'open'
        ORDER BY toner_requirement.created_at DESC
        LIMIT 30
    """).fetchall()

    return render_template(
        'toner_management.html', user=user, entries=entries, locations=LOCATIONS, departments=DEPARTMENTS,
        toner_types=TONER_TYPES, today=selected_date.isoformat(), active_page='toner_management',
        pending_followups=pending_followups, open_reqs=open_reqs,
    )


@app.route('/toner/<int:entry_id>/set-reached', methods=['POST'])
@login_required
def set_toner_reached(entry_id):
    user = current_user()
    db = get_db()
    e = db.execute("SELECT * FROM toner_entry WHERE id = ?", (entry_id,)).fetchone()
    if not e or (e['user_id'] != user['id'] and user['role'] != 'admin'):
        flash('Toner entry not found.', 'error')
        return redirect(url_for('toner_management'))

    if not reached_eligible(e['created_at']):
        flash('Reached follow-up is available only after 1 day of submission.', 'error')
        return redirect(url_for('toner_management'))

    value = request.form.get('reached_value')
    if value not in ('Yes', 'No'):
        flash('Invalid selection.', 'error')
        return redirect(url_for('toner_management'))

    db.execute("UPDATE toner_entry SET reached = ? WHERE id = ?", (value, entry_id))
    db.commit()
    flash('Toner follow-up (Reached) saved.', 'success')
    return redirect(url_for('toner_management'))


@app.route('/toner/<int:entry_id>/edit', methods=['POST'])
@login_required
def edit_toner_entry(entry_id):
    user = current_user()
    db = get_db()
    e = db.execute("SELECT * FROM toner_entry WHERE id = ?", (entry_id,)).fetchone()
    redirect_target = url_for('admin_toner') if user['role'] == 'admin' else url_for('toner_management')

    if not e or (e['user_id'] != user['id'] and user['role'] != 'admin'):
        flash('Toner entry not found.', 'error')
        return redirect(redirect_target)

    entry_date = parse_date(request.form.get('entry_date'))
    toner_type = request.form.get('toner_type', '').strip()
    if toner_type == 'Other':
        toner_type = request.form.get('toner_type_other', '').strip() or 'Other'
    try:
        qty = int(request.form.get('quantity') or 1)
    except ValueError:
        qty = 1
    if qty < 1:
        qty = 1

    db.execute(
        "UPDATE toner_entry SET entry_date=?, location=?, department=?, handover_to=?, toner_type=?, quantity=? WHERE id=?",
        (entry_date.isoformat(), request.form.get('location', '').strip(),
         request.form.get('department', '').strip(), request.form.get('handover_to', '').strip(),
         toner_type, qty, entry_id)
    )
    db.commit()
    flash('Toner entry updated.', 'success')
    return redirect(redirect_target)


@app.route('/toner/<int:entry_id>/delete', methods=['POST'])
@login_required
def delete_toner_entry(entry_id):
    user = current_user()
    db = get_db()
    e = db.execute("SELECT * FROM toner_entry WHERE id = ?", (entry_id,)).fetchone()
    redirect_target = url_for('admin_toner') if user['role'] == 'admin' else url_for('toner_management')
    if not e or (e['user_id'] != user['id'] and user['role'] != 'admin'):
        flash('Toner entry not found or not allowed.', 'error')
        return redirect(redirect_target)
    db.execute("DELETE FROM toner_entry WHERE id = ?", (entry_id,))
    db.commit()
    flash('Toner entry deleted.', 'success')
    return redirect(redirect_target)


@app.route('/admin/toner')
@admin_required
def admin_toner():
    db = get_db()
    rows = db.execute("""
        SELECT toner_entry.*, user.full_name AS employee_name
        FROM toner_entry JOIN user ON toner_entry.user_id = user.id
        ORDER BY toner_entry.entry_date DESC, toner_entry.id DESC
    """).fetchall()
    open_reqs = db.execute("""
        SELECT toner_requirement.*, user.full_name AS employee_name
        FROM toner_requirement JOIN user ON toner_requirement.user_id = user.id
        WHERE COALESCE(toner_requirement.status, 'open') = 'open'
        ORDER BY toner_requirement.created_at DESC
        LIMIT 30
    """).fetchall()
    return render_template(
        'admin_toner.html', user=current_user(), entries=rows, locations=LOCATIONS, departments=DEPARTMENTS,
        toner_types=TONER_TYPES, active_page='admin_toner', open_reqs=open_reqs,
    )


@app.route('/toner-requirement/<int:entry_id>/ack', methods=['POST'])
@login_required
def ack_toner_requirement(entry_id):
    """Mark a toner requirement as seen/handled from the Toner Management popup."""
    user = current_user()
    if user['role'] == 'toner_req':
        return redirect(url_for('toner_req_portal'))
    db = get_db()
    e = db.execute("SELECT * FROM toner_requirement WHERE id = ?", (entry_id,)).fetchone()
    if not e:
        flash('Requirement not found.', 'error')
    else:
        db.execute("UPDATE toner_requirement SET status = 'seen' WHERE id = ?", (entry_id,))
        db.commit()
        flash('Toner requirement acknowledged.', 'success')
    if user['role'] == 'admin':
        return redirect(url_for('admin_toner'))
    return redirect(url_for('toner_management'))


# ---------------------------------------------------------------------------
# Routes: Toner Requirement (visible to all logged-in users)
# ---------------------------------------------------------------------------

@app.route('/toner-requirement', methods=['GET', 'POST'])
@login_required
def toner_requirement():
    user = current_user()
    # Dedicated portal users should stay on the simple interface
    if user and user['role'] == 'toner_req':
        return redirect(url_for('toner_req_portal'))

    db = get_db()

    if request.method == 'POST':
        entry_date = parse_date(request.form.get('entry_date'))
        toner_type = request.form.get('toner_type', '').strip()
        if toner_type == 'Other':
            toner_type = request.form.get('toner_type_other', '').strip() or 'Other'
        try:
            qty = int(request.form.get('quantity') or 1)
        except ValueError:
            qty = 1
        if qty < 1:
            qty = 1

        db.execute(
            "INSERT INTO toner_requirement (user_id, entry_date, requester_name, needed_by, department, location, toner_type, quantity, remark, status, created_at) "
            "VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, 'open', ?)",
            (user['id'], entry_date.isoformat(),
             request.form.get('requester_name', '').strip(),
             request.form.get('needed_by', '').strip(),
             request.form.get('department', '').strip(),
             request.form.get('location', '').strip(),
             toner_type, qty,
             request.form.get('remark', '').strip(),
             datetime.utcnow().isoformat())
        )
        db.commit()
        flash('Toner requirement submitted. Team will see a popup in Toner Management.', 'success')
        return redirect(url_for('toner_requirement', date=entry_date.isoformat()))

    selected_date = parse_date(request.args.get('date')) if request.args.get('date') else date.today()
    rows = db.execute("""
        SELECT toner_requirement.*, user.full_name AS employee_name
        FROM toner_requirement JOIN user ON toner_requirement.user_id = user.id
        ORDER BY toner_requirement.entry_date DESC, toner_requirement.id DESC
    """).fetchall()

    return render_template(
        'toner_requirement.html',
        user=user,
        entries=rows,
        locations=LOCATIONS,
        departments=DEPARTMENTS,
        toner_types=TONER_TYPES,
        today=selected_date.isoformat(),
        active_page='toner_requirement',
    )


@app.route('/toner-req-portal', methods=['GET', 'POST'])
@login_required
def toner_req_portal():
    """Minimal interface for Toner Requirement-only users (department link accounts)."""
    user = current_user()
    if not user or user['role'] != 'toner_req':
        flash('This page is only for Toner Requirement accounts.', 'error')
        return redirect(home_for_role(session.get('role')))

    db = get_db()
    if request.method == 'POST':
        entry_date = parse_date(request.form.get('entry_date'))
        toner_type = request.form.get('toner_type', '').strip()
        if toner_type == 'Other':
            toner_type = request.form.get('toner_type_other', '').strip() or 'Other'
        try:
            qty = int(request.form.get('quantity') or 1)
        except ValueError:
            qty = 1
        if qty < 1:
            qty = 1

        db.execute(
            "INSERT INTO toner_requirement (user_id, entry_date, requester_name, needed_by, department, location, toner_type, quantity, remark, status, created_at) "
            "VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, 'open', ?)",
            (user['id'], entry_date.isoformat(),
             request.form.get('requester_name', '').strip(),
             request.form.get('needed_by', '').strip(),
             request.form.get('department', '').strip(),
             request.form.get('location', '').strip(),
             toner_type, qty,
             request.form.get('remark', '').strip(),
             datetime.utcnow().isoformat())
        )
        db.commit()
        flash('Toner requirement submitted successfully. IT team will see it in Toner Management.', 'success')
        return redirect(url_for('toner_req_portal'))

    my_entries = db.execute(
        "SELECT * FROM toner_requirement WHERE user_id = ? ORDER BY entry_date DESC, id DESC LIMIT 20",
        (user['id'],)
    ).fetchall()

    return render_template(
        'toner_req_portal.html',
        user=user,
        toner_types=TONER_TYPES,
        departments=DEPARTMENTS,
        today=date.today().isoformat(),
        entries=my_entries,
    )


@app.route('/toner-requirement/<int:entry_id>/delete', methods=['POST'])
@login_required
def delete_toner_requirement(entry_id):
    user = current_user()
    db = get_db()
    e = db.execute("SELECT * FROM toner_requirement WHERE id = ?", (entry_id,)).fetchone()
    if not e:
        flash('Requirement not found.', 'error')
        return redirect(url_for('toner_requirement'))
    # Owner or admin can delete
    if e['user_id'] != user['id'] and user['role'] != 'admin':
        flash('You can only delete your own requirements.', 'error')
        return redirect(url_for('toner_requirement'))
    db.execute("DELETE FROM toner_requirement WHERE id = ?", (entry_id,))
    db.commit()
    flash('Toner requirement deleted.', 'success')
    return redirect(url_for('toner_requirement'))


# ---------------------------------------------------------------------------
# Routes: CCTV Register
# ---------------------------------------------------------------------------

@app.route('/cctv', methods=['GET', 'POST'])
@login_required
def cctv_register():
    user = current_user()
    db = get_db()

    if request.method == 'POST':
        entry_date = parse_date(request.form.get('entry_date'))
        db.execute(
            "INSERT INTO cctv_entry (user_id, entry_date, viewer_name, department, location, purpose, remark, created_at) "
            "VALUES (?, ?, ?, ?, ?, ?, ?, ?)",
            (user['id'], entry_date.isoformat(), request.form.get('viewer_name', '').strip(),
             request.form.get('department', '').strip(), request.form.get('location', '').strip(),
             request.form.get('purpose', '').strip() or DEFAULT_CCTV_PURPOSE,
             request.form.get('remark', '').strip(), datetime.utcnow().isoformat())
        )
        db.commit()
        flash('CCTV register entry added.', 'success')
        return redirect(url_for('cctv_register', date=entry_date.isoformat()))

    selected_date = parse_date(request.args.get('date')) if request.args.get('date') else date.today()
    entries = db.execute(
        "SELECT * FROM cctv_entry WHERE user_id = ? ORDER BY entry_date DESC, id DESC", (user['id'],)
    ).fetchall()
    return render_template(
        'cctv_register.html', user=user, entries=entries, locations=LOCATIONS, departments=DEPARTMENTS,
        today=selected_date.isoformat(), active_page='cctv_register',
    )


@app.route('/cctv/<int:entry_id>/edit', methods=['POST'])
@login_required
def edit_cctv_entry(entry_id):
    user = current_user()
    db = get_db()
    e = db.execute("SELECT * FROM cctv_entry WHERE id = ?", (entry_id,)).fetchone()
    redirect_target = url_for('admin_cctv') if user['role'] == 'admin' else url_for('cctv_register')

    if not e or (e['user_id'] != user['id'] and user['role'] != 'admin'):
        flash('CCTV register entry not found.', 'error')
        return redirect(redirect_target)

    entry_date = parse_date(request.form.get('entry_date'))
    db.execute(
        "UPDATE cctv_entry SET entry_date=?, viewer_name=?, department=?, location=?, purpose=?, remark=? WHERE id=?",
        (entry_date.isoformat(), request.form.get('viewer_name', '').strip(),
         request.form.get('department', '').strip(), request.form.get('location', '').strip(),
         request.form.get('purpose', '').strip() or DEFAULT_CCTV_PURPOSE,
         request.form.get('remark', '').strip(), entry_id)
    )
    db.commit()
    flash('CCTV register entry updated.', 'success')
    return redirect(redirect_target)


@app.route('/cctv/<int:entry_id>/delete', methods=['POST'])
@login_required
def delete_cctv_entry(entry_id):
    user = current_user()
    db = get_db()
    e = db.execute("SELECT * FROM cctv_entry WHERE id = ?", (entry_id,)).fetchone()
    redirect_target = url_for('admin_cctv') if user['role'] == 'admin' else url_for('cctv_register')
    if not e or (e['user_id'] != user['id'] and user['role'] != 'admin'):
        flash('CCTV entry not found or not allowed.', 'error')
        return redirect(redirect_target)
    db.execute("DELETE FROM cctv_entry WHERE id = ?", (entry_id,))
    db.commit()
    flash('CCTV entry deleted.', 'success')
    return redirect(redirect_target)


@app.route('/admin/cctv')
@admin_required
def admin_cctv():
    db = get_db()
    rows = db.execute("""
        SELECT cctv_entry.*, user.full_name AS employee_name
        FROM cctv_entry JOIN user ON cctv_entry.user_id = user.id
        ORDER BY cctv_entry.entry_date DESC, cctv_entry.id DESC
    """).fetchall()
    return render_template(
        'admin_cctv.html', user=current_user(), entries=rows, locations=LOCATIONS, departments=DEPARTMENTS,
        active_page='admin_cctv',
    )


# ---------------------------------------------------------------------------
# Routes: Visitor Register
# ---------------------------------------------------------------------------

@app.route('/visitor-register', methods=['GET', 'POST'])
@login_required
def visitor_register():
    user = current_user()
    if user['role'] == 'toner_req':
        return redirect(url_for('toner_req_portal'))
    db = get_db()

    if request.method == 'POST':
        entry_date = parse_date(request.form.get('entry_date'))
        db.execute(
            "INSERT INTO visitor_entry (user_id, entry_date, location, visitor_name, purpose, phone, created_at) "
            "VALUES (?, ?, ?, ?, ?, ?, ?)",
            (user['id'], entry_date.isoformat(),
             request.form.get('location', '').strip(),
             request.form.get('visitor_name', '').strip(),
             request.form.get('purpose', '').strip(),
             request.form.get('phone', '').strip(),
             datetime.utcnow().isoformat())
        )
        db.commit()
        flash('Visitor entry added.', 'success')
        return redirect(url_for('visitor_register', date=entry_date.isoformat()))

    selected_date = parse_date(request.args.get('date')) if request.args.get('date') else date.today()
    entries = db.execute(
        "SELECT * FROM visitor_entry WHERE user_id = ? ORDER BY entry_date DESC, id DESC", (user['id'],)
    ).fetchall()
    return render_template(
        'visitor_register.html', user=user, entries=entries, locations=LOCATIONS,
        today=selected_date.isoformat(), active_page='visitor_register',
    )


@app.route('/visitor-register/<int:entry_id>/delete', methods=['POST'])
@login_required
def delete_visitor_entry(entry_id):
    user = current_user()
    db = get_db()
    e = db.execute("SELECT * FROM visitor_entry WHERE id = ?", (entry_id,)).fetchone()
    redirect_target = url_for('admin_visitor') if user['role'] == 'admin' else url_for('visitor_register')
    if not e or (e['user_id'] != user['id'] and user['role'] != 'admin'):
        flash('Visitor entry not found.', 'error')
        return redirect(redirect_target)
    db.execute("DELETE FROM visitor_entry WHERE id = ?", (entry_id,))
    db.commit()
    flash('Visitor entry deleted.', 'success')
    return redirect(redirect_target)


@app.route('/admin/visitor-register')
@admin_required
def admin_visitor():
    db = get_db()
    rows = db.execute("""
        SELECT visitor_entry.*, user.full_name AS employee_name
        FROM visitor_entry JOIN user ON visitor_entry.user_id = user.id
        ORDER BY visitor_entry.entry_date DESC, visitor_entry.id DESC
    """).fetchall()
    return render_template(
        'admin_visitor.html', user=current_user(), entries=rows, locations=LOCATIONS,
        active_page='admin_visitor',
    )


# ---------------------------------------------------------------------------
# Routes: Admin
# ---------------------------------------------------------------------------

@app.route('/admin')
@admin_required
def admin_dashboard():
    db = get_db()
    total_users = db.execute("SELECT COUNT(*) c FROM user").fetchone()['c']
    total_reports = db.execute("SELECT COUNT(*) c FROM report").fetchone()['c']
    pending_reports = db.execute("SELECT COUNT(*) c FROM report WHERE status = 'Pending'").fetchone()['c']
    completed_reports = db.execute("SELECT COUNT(*) c FROM report WHERE status = 'Completed'").fetchone()['c']
    admin_count = db.execute("SELECT COUNT(*) c FROM user WHERE role = 'admin'").fetchone()['c']
    users = db.execute("SELECT * FROM user ORDER BY id ASC").fetchall()
    return render_template(
        'admin_dashboard.html',
        user=current_user(),
        users=users,
        total_users=total_users,
        total_reports=total_reports,
        pending_reports=pending_reports,
        completed_reports=completed_reports,
        admin_count=admin_count,
        departments=DEPARTMENTS,
        active_page='admin_dashboard',
    )


@app.route('/admin/reports')
@admin_required
def admin_reports():
    db = get_db()
    rows = db.execute("""
        SELECT report.*, user.full_name AS employee_name
        FROM report JOIN user ON report.user_id = user.id
        ORDER BY report.report_date DESC
    """).fetchall()
    reports = []
    for r in rows:
        d = dict(r)
        d['reached_eligible'] = reached_eligible(r['created_at'])
        reports.append(d)
    return render_template('all_reports.html', user=current_user(), reports=reports,
                            locations=LOCATIONS, departments=DEPARTMENTS, active_page='admin_reports')


@app.route('/admin/reports/<int:report_id>/edit', methods=['POST'])
@admin_required
def edit_report(report_id):
    db = get_db()
    r = db.execute("SELECT * FROM report WHERE id = ?", (report_id,)).fetchone()
    if not r:
        flash('Report not found.', 'error')
        return redirect(url_for('admin_reports'))

    report_date = parse_date(request.form.get('report_date'))
    db.execute(
        "UPDATE report SET report_date=?, customer_name=?, location=?, department=?, mode=?, status=?, issue=? "
        "WHERE id=?",
        (report_date.isoformat(), request.form.get('customer_name', '').strip(),
         request.form.get('location', '').strip(), request.form.get('department', '').strip(),
         request.form.get('mode'), request.form.get('status'), request.form.get('issue'), report_id)
    )
    db.commit()
    flash('Report updated successfully — the mistake has been corrected.', 'success')
    return redirect(url_for('admin_reports'))


@app.route('/admin/users/create', methods=['POST'])
@admin_required
def create_user():
    db = get_db()
    username = request.form.get('username', '').strip()
    full_name = request.form.get('full_name', '').strip()
    email = request.form.get('email', '').strip()
    department = request.form.get('department', '')
    role = request.form.get('role', 'employee')
    password = request.form.get('password', '') or 'welcome123'

    if role not in ('employee', 'admin', 'toner_req'):
        role = 'employee'

    existing = db.execute("SELECT id FROM user WHERE username = ?", (username,)).fetchone()
    if existing:
        flash('Username already exists.', 'error')
        return redirect(url_for('admin_dashboard'))

    db.execute(
        "INSERT INTO user (username, full_name, email, password_hash, role, department, is_active, created_at) "
        "VALUES (?, ?, ?, ?, ?, ?, 1, ?)",
        (username, full_name, email, generate_password_hash(password), role, department,
         datetime.utcnow().isoformat())
    )
    db.commit()
    if role == 'toner_req':
        flash(
            f'Toner Requirement user "{username}" created (password: "{password}"). '
            f'They will only see the toner request form after login.',
            'success',
        )
    else:
        flash(f'User "{username}" created with temporary password "{password}".', 'success')
    return redirect(url_for('admin_dashboard'))


@app.route('/admin/users/<int:user_id>/reset-password', methods=['POST'])
@admin_required
def reset_password(user_id):
    db = get_db()
    u = db.execute("SELECT * FROM user WHERE id = ?", (user_id,)).fetchone()
    if not u:
        flash('User not found.', 'error')
        return redirect(url_for('admin_dashboard'))
    new_password = request.form.get('new_password') or 'welcome123'
    db.execute("UPDATE user SET password_hash = ? WHERE id = ?", (generate_password_hash(new_password), user_id))
    db.commit()
    flash(f'Password for "{u["username"]}" reset successfully.', 'success')
    return redirect(url_for('admin_dashboard'))


@app.route('/admin/users/<int:user_id>/toggle-active', methods=['POST'])
@admin_required
def toggle_active(user_id):
    db = get_db()
    u = db.execute("SELECT * FROM user WHERE id = ?", (user_id,)).fetchone()
    if not u:
        flash('User not found.', 'error')
        return redirect(url_for('admin_dashboard'))
    if u['role'] == 'admin' and u['id'] == session.get('user_id'):
        flash('You cannot deactivate your own admin account.', 'error')
        return redirect(url_for('admin_dashboard'))
    new_status = 0 if u['is_active'] else 1
    db.execute("UPDATE user SET is_active = ? WHERE id = ?", (new_status, user_id))
    db.commit()
    flash(f'User "{u["username"]}" is now {"active" if new_status else "inactive"}.', 'success')
    return redirect(url_for('admin_dashboard'))


@app.route('/admin/users/<int:user_id>/delete', methods=['POST'])
@admin_required
def delete_user(user_id):
    if user_id == session.get('user_id'):
        flash('You cannot delete your own account while logged in.', 'error')
        return redirect(url_for('admin_dashboard'))
    db = get_db()
    u = db.execute("SELECT * FROM user WHERE id = ?", (user_id,)).fetchone()
    if not u:
        flash('User not found.', 'error')
        return redirect(url_for('admin_dashboard'))
    db.execute("DELETE FROM report WHERE user_id = ?", (user_id,))
    db.execute("DELETE FROM user WHERE id = ?", (user_id,))
    db.commit()
    flash(f'User "{u["username"]}" deleted.', 'success')
    return redirect(url_for('admin_dashboard'))


# ---------------------------------------------------------------------------
# Export helpers
# ---------------------------------------------------------------------------

def get_reports_for_range(range_type, ref_date, user_id=None, from_date=None, to_date=None):
    db = get_db()
    base = """
        SELECT report.*, user.full_name AS employee_name
        FROM report JOIN user ON report.user_id = user.id
        WHERE 1=1
    """
    params = []
    if user_id:
        base += " AND report.user_id = ?"
        params.append(user_id)

    if range_type == 'daily':
        base += " AND report.report_date = ?"
        params.append(ref_date.isoformat())
    elif range_type == 'monthly':
        base += " AND strftime('%Y-%m', report.report_date) = ?"
        params.append(ref_date.strftime('%Y-%m'))
    elif range_type == 'yearly':
        base += " AND strftime('%Y', report.report_date) = ?"
        params.append(str(ref_date.year))
    elif range_type == 'custom':
        fd = from_date or ref_date
        td = to_date or ref_date
        if fd > td:
            fd, td = td, fd
        base += " AND report.report_date >= ? AND report.report_date <= ?"
        params.append(fd.isoformat())
        params.append(td.isoformat())

    base += " ORDER BY report.report_date ASC"
    return db.execute(base, params).fetchall()


def build_excel(reports, title, include_employee=False):
    """Build call-report Excel.
    include_employee=True only for admin 'all' exports (shows who submitted).
    Own exports omit Employee so the first column is clearly the report Date.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Call Reports"

    if include_employee:
        headers = ["Date", "Employee", "User Name", "Location", "Department", "Mode", "Status", "Issue / Description"]
        widths = [14, 20, 20, 26, 18, 12, 12, 45]
    else:
        headers = ["Date", "User Name", "Location", "Department", "Mode", "Status", "Issue / Description"]
        widths = [14, 20, 26, 18, 12, 12, 45]
    ws.append(headers)
    header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for r in reports:
        d = datetime.strptime(r['report_date'], '%Y-%m-%d').strftime('%d %b %Y')
        entry_type = r['entry_type'] if 'entry_type' in r.keys() else 'Visit'
        if entry_type == 'Holiday':
            row = [d]
            if include_employee:
                row.append(r['employee_name'] or '')
            row.extend(['-', 'HOLIDAY', '-', '-', 'Holiday', 'Weekly Holiday / Non-working day'])
            ws.append(row)
        elif entry_type == 'Leave':
            leave_label = r['leave_type'] if ('leave_type' in r.keys() and r['leave_type']) else 'Leave'
            row = [d]
            if include_employee:
                row.append(r['employee_name'] or '')
            row.extend(['-', 'ON LEAVE', '-', '-', 'Leave', leave_label])
            ws.append(row)
        else:
            row = [d]
            if include_employee:
                row.append(r['employee_name'] or '')
            row.extend([
                r['customer_name'] or '', r['location'] or '', r['department'] or '',
                r['mode'] or '', r['status'] or '', r['issue'] or '',
            ])
            ws.append(row)

    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def _pdf_footer(canvas, doc):
    canvas.saveState()
    canvas.setFont('Helvetica', 8)
    canvas.setFillColor(colors.HexColor('#64748b'))
    page_width = doc.pagesize[0]
    canvas.drawCentredString(page_width / 2, 10 * mm, "Developed by AZAAN")
    canvas.restoreState()


def build_pdf(reports, title, include_employee=False):
    from reportlab.lib.styles import ParagraphStyle
    buf = io.BytesIO()
    # Narrow side margins so Date column has enough room on landscape A4
    doc = SimpleDocTemplate(
        buf, pagesize=landscape(A4),
        topMargin=14 * mm, bottomMargin=16 * mm,
        leftMargin=10 * mm, rightMargin=10 * mm,
    )
    styles = getSampleStyleSheet()
    cell_style = ParagraphStyle(
        'CellSmall', parent=styles['Normal'],
        fontSize=7, leading=9, wordWrap='CJK',
    )
    header_style = ParagraphStyle(
        'HeaderSmall', parent=styles['Normal'],
        fontSize=8, leading=10, textColor=colors.white, fontName='Helvetica-Bold',
    )
    elements = [Paragraph(title, styles['Title']), Spacer(1, 6)]

    def _p(text):
        """Wrap cell text so long Location/Issue values wrap instead of crushing Date."""
        safe = (str(text) if text is not None else '').replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;')
        return Paragraph(safe, cell_style)

    def _h(text):
        return Paragraph(str(text), header_style)

    if include_employee:
        headers = ["Date", "Employee", "User Name", "Location", "Department", "Mode", "Status", "Issue / Description"]
        # Total ~ usable width on landscape A4 with 10mm margins ≈ 277mm ≈ 785pt
        col_widths = [72, 70, 70, 130, 65, 45, 55, 278]
    else:
        headers = ["Date", "User Name", "Location", "Department", "Mode", "Status", "Issue / Description"]
        col_widths = [78, 85, 150, 70, 48, 58, 296]

    data = [[_h(h) for h in headers]]
    holiday_rows = []
    leave_rows = []
    for idx, r in enumerate(reports, start=1):
        d = datetime.strptime(r['report_date'], '%Y-%m-%d').strftime('%d %b %Y')
        entry_type = r['entry_type'] if 'entry_type' in r.keys() else 'Visit'
        if entry_type == 'Holiday':
            row = [_p(d)]
            if include_employee:
                row.append(_p(r['employee_name'] or ''))
            row.extend([_p('-'), _p('HOLIDAY'), _p('-'), _p('-'), _p('Holiday'), _p('Weekly Holiday / Non-working day')])
            data.append(row)
            holiday_rows.append(idx)
        elif entry_type == 'Leave':
            leave_label = r['leave_type'] if ('leave_type' in r.keys() and r['leave_type']) else 'Leave'
            row = [_p(d)]
            if include_employee:
                row.append(_p(r['employee_name'] or ''))
            row.extend([_p('-'), _p('ON LEAVE'), _p('-'), _p('-'), _p('Leave'), _p(leave_label)])
            data.append(row)
            leave_rows.append(idx)
        else:
            row = [_p(d)]
            if include_employee:
                row.append(_p(r['employee_name'] or ''))
            row.extend([
                _p(r['customer_name'] or ''),
                _p(r['location'] or ''),
                _p(r['department'] or ''),
                _p(r['mode'] or ''),
                _p(r['status'] or ''),
                _p((r['issue'] or '')[:200]),
            ])
            data.append(row)

    table_style = [
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1E3A8A')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTSIZE', (0, 0), (-1, -1), 7),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.whitesmoke, colors.white]),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('LEFTPADDING', (0, 0), (-1, -1), 3),
        ('RIGHTPADDING', (0, 0), (-1, -1), 3),
        ('TOPPADDING', (0, 0), (-1, -1), 3),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
        # Keep Date column readable
        ('FONTNAME', (0, 1), (0, -1), 'Helvetica-Bold'),
    ]
    for ridx in holiday_rows:
        table_style.append(('BACKGROUND', (0, ridx), (-1, ridx), colors.HexColor('#DBEAFE')))
        table_style.append(('TEXTCOLOR', (0, ridx), (-1, ridx), colors.HexColor('#1E3A8A')))
    for ridx in leave_rows:
        table_style.append(('BACKGROUND', (0, ridx), (-1, ridx), colors.HexColor('#FEF3C7')))
        table_style.append(('TEXTCOLOR', (0, ridx), (-1, ridx), colors.HexColor('#92400E')))

    table = Table(data, colWidths=col_widths, repeatRows=1)
    table.setStyle(TableStyle(table_style))
    elements.append(table)
    doc.build(elements, onFirstPage=_pdf_footer, onLaterPages=_pdf_footer)
    buf.seek(0)
    return buf


def build_petrol_excel(entries, title, total_amount):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Petrol Expenses"

    # Export columns: Date, Employee, Location, Amount only (no Reported At / KM)
    headers = ["Date", "Employee", "Location Visited", "Amount (Rs.)"]
    ws.append(headers)
    header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
    sunday_fill = PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for e in entries:
        d = datetime.strptime(e['expense_date'], '%Y-%m-%d').strftime('%d %b %Y')
        name = e['employee_name'] if 'employee_name' in e.keys() else ''
        day_type = e['day_type'] if ('day_type' in e.keys() and e['day_type']) else 'Working'
        if day_type == 'Sunday':
            ws.append([d, name, 'SUNDAY — Weekly Holiday / Non-working day', 0])
            for cell in ws[ws.max_row]:
                cell.fill = sunday_fill
        else:
            ws.append([d, name, e['location'] or '', round(e['amount'] or 0, 2)])

    ws.append(['', '', 'Total', round(total_amount, 2)])
    for cell in ws[ws.max_row]:
        cell.font = Font(bold=True)

    widths = [14, 20, 40, 14]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def build_petrol_pdf(entries, title, total_amount):
    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=A4,
        topMargin=14 * mm, bottomMargin=16 * mm,
        leftMargin=12 * mm, rightMargin=12 * mm,
    )
    styles = getSampleStyleSheet()
    elements = [Paragraph(title, styles['Title']), Spacer(1, 8)]

    # Export columns: Date, Employee, Location, Amount only (no Reported At / KM)
    data = [["Date", "Employee", "Location Visited", "Amount (Rs.)"]]
    sunday_rows = []
    for idx, e in enumerate(entries, start=1):
        d = datetime.strptime(e['expense_date'], '%Y-%m-%d').strftime('%d %b %Y')
        name = e['employee_name'] if 'employee_name' in e.keys() else ''
        day_type = e['day_type'] if ('day_type' in e.keys() and e['day_type']) else 'Working'
        if day_type == 'Sunday':
            data.append([d, name, 'SUNDAY — Weekly Holiday / Non-working day', '0.00'])
            sunday_rows.append(idx)
        else:
            data.append([d, name, e['location'] or '', f"{e['amount'] or 0:.2f}"])
    data.append(["", "", "Total", f"{total_amount:.2f}"])

    table_style = [
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1E3A8A')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -2), [colors.whitesmoke, colors.white]),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
    ]
    for ridx in sunday_rows:
        table_style.append(('BACKGROUND', (0, ridx), (-1, ridx), colors.HexColor('#DBEAFE')))
        table_style.append(('TEXTCOLOR', (0, ridx), (-1, ridx), colors.HexColor('#1E3A8A')))

    table = Table(data, colWidths=[80, 100, 220, 80], repeatRows=1)
    table.setStyle(TableStyle(table_style))
    elements.append(table)
    doc.build(elements, onFirstPage=_pdf_footer, onLaterPages=_pdf_footer)
    buf.seek(0)
    return buf


@app.route('/export/<range_type>/<fmt>')
@login_required
def export_reports(range_type, fmt):
    if range_type not in ('daily', 'monthly', 'yearly', 'custom'):
        flash('Invalid export range.', 'error')
        return redirect(url_for('index'))

    ref_date = parse_date(request.args.get('date'))
    from_date = parse_date(request.args.get('from')) if request.args.get('from') else None
    to_date = parse_date(request.args.get('to')) if request.args.get('to') else None
    user = current_user()
    scope_all = user['role'] == 'admin' and request.args.get('scope') == 'all'
    scope_user_id = None if scope_all else user['id']

    reports = get_reports_for_range(
        range_type, ref_date, user_id=scope_user_id,
        from_date=from_date, to_date=to_date,
    )

    if range_type == 'custom':
        fd = from_date or ref_date
        td = to_date or ref_date
        if fd > td:
            fd, td = td, fd
        label = f"{fd.strftime('%d %b %Y')} to {td.strftime('%d %b %Y')}"
        filename_base = f"call_report_custom_{fd.isoformat()}_to_{td.isoformat()}"
    else:
        label = {'daily': ref_date.strftime('%d %b %Y'),
                 'monthly': ref_date.strftime('%B %Y'),
                 'yearly': str(ref_date.year)}[range_type]
        filename_base = f"call_report_{range_type}_{ref_date.isoformat()}"
    title = f"Call Report - {label}"

    # Admin exporting everyone → keep Employee column; own export → Date first, no self-name column
    include_employee = scope_all

    if fmt == 'excel':
        buf = build_excel(reports, title, include_employee=include_employee)
        return send_file(buf, as_attachment=True, download_name=f"{filename_base}.xlsx",
                          mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    elif fmt == 'pdf':
        buf = build_pdf(reports, title, include_employee=include_employee)
        return send_file(buf, as_attachment=True, download_name=f"{filename_base}.pdf", mimetype='application/pdf')
    else:
        flash('Invalid export format.', 'error')
        return redirect(url_for('index'))


def get_petrol_for_range(range_type, ref_date, user_id=None):
    db = get_db()
    base = """
        SELECT petrol_expense.*, user.full_name AS employee_name
        FROM petrol_expense JOIN user ON petrol_expense.user_id = user.id
        WHERE 1=1
    """
    params = []
    if user_id:
        base += " AND petrol_expense.user_id = ?"
        params.append(user_id)

    if range_type == 'daily':
        base += " AND petrol_expense.expense_date = ?"
        params.append(ref_date.isoformat())
    elif range_type == 'monthly':
        base += " AND strftime('%Y-%m', petrol_expense.expense_date) = ?"
        params.append(ref_date.strftime('%Y-%m'))
    elif range_type == 'yearly':
        base += " AND strftime('%Y', petrol_expense.expense_date) = ?"
        params.append(str(ref_date.year))

    base += " ORDER BY petrol_expense.expense_date ASC"
    return db.execute(base, params).fetchall()


@app.route('/petrol-expenses/export/<range_type>/<fmt>')
@login_required
def export_petrol_expenses(range_type, fmt):
    if range_type not in ('daily', 'monthly', 'yearly'):
        flash('Invalid export range.', 'error')
        return redirect(url_for('petrol_expenses'))

    ref_date = parse_date(request.args.get('date'))
    user = current_user()
    scope_user_id = None if (user['role'] == 'admin' and request.args.get('scope') == 'all') else user['id']

    entries = get_petrol_for_range(range_type, ref_date, user_id=scope_user_id)
    total_amount = sum(e['amount'] or 0 for e in entries)

    label = {'daily': ref_date.strftime('%d %b %Y'),
             'monthly': ref_date.strftime('%B %Y'),
             'yearly': str(ref_date.year)}[range_type]
    title = f"Petrol Expenses - {label}"
    filename_base = f"petrol_expenses_{range_type}_{ref_date.isoformat()}"

    if fmt == 'excel':
        buf = build_petrol_excel(entries, title, total_amount)
        return send_file(buf, as_attachment=True, download_name=f"{filename_base}.xlsx",
                          mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    elif fmt == 'pdf':
        buf = build_petrol_pdf(entries, title, total_amount)
        return send_file(buf, as_attachment=True, download_name=f"{filename_base}.pdf", mimetype='application/pdf')
    else:
        flash('Invalid export format.', 'error')
        return redirect(url_for('petrol_expenses'))


# ---------------------------------------------------------------------------
# Generic table export (Excel + PDF) for Toner / Requirement / CCTV / Visitor
# ---------------------------------------------------------------------------

def _range_label(range_type, ref_date):
    return {
        'daily': ref_date.strftime('%d %b %Y'),
        'monthly': ref_date.strftime('%B %Y'),
        'yearly': str(ref_date.year),
    }[range_type]


def _apply_date_filter(sql, params, date_col, range_type, ref_date):
    if range_type == 'daily':
        sql += f" AND {date_col} = ?"
        params.append(ref_date.isoformat())
    elif range_type == 'monthly':
        sql += f" AND strftime('%Y-%m', {date_col}) = ?"
        params.append(ref_date.strftime('%Y-%m'))
    elif range_type == 'yearly':
        sql += f" AND strftime('%Y', {date_col}) = ?"
        params.append(str(ref_date.year))
    return sql, params


def build_generic_excel(headers, rows, sheet_title="Export"):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_title[:31]
    ws.append(headers)
    header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    for row in rows:
        ws.append(list(row))
    for i in range(1, len(headers) + 1):
        col_letter = openpyxl.utils.get_column_letter(i)
        ws.column_dimensions[col_letter].width = 16 if i < len(headers) else 28
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def build_generic_pdf(title, headers, rows, landscape_mode=True):
    buf = io.BytesIO()
    pagesize = landscape(A4) if landscape_mode else A4
    doc = SimpleDocTemplate(buf, pagesize=pagesize, topMargin=18 * mm, bottomMargin=18 * mm)
    styles = getSampleStyleSheet()
    elements = [Paragraph(title, styles['Title']), Spacer(1, 8)]
    data = [list(headers)]
    for row in rows:
        data.append([str(c) if c is not None else '' for c in row])
    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1E3A8A')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTSIZE', (0, 0), (-1, -1), 7),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.whitesmoke, colors.white]),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
    ]))
    elements.append(table)
    doc.build(elements, onFirstPage=_pdf_footer, onLaterPages=_pdf_footer)
    buf.seek(0)
    return buf


def _send_export(buf, fmt, filename_base):
    if fmt == 'excel':
        return send_file(buf, as_attachment=True, download_name=f"{filename_base}.xlsx",
                          mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    return send_file(buf, as_attachment=True, download_name=f"{filename_base}.pdf", mimetype='application/pdf')


def _fmt_row_date(val):
    if not val:
        return ''
    try:
        return datetime.strptime(val, '%Y-%m-%d').strftime('%d %b %Y')
    except (ValueError, TypeError):
        return str(val)


@app.route('/toner/export/<range_type>/<fmt>')
@login_required
def export_toner(range_type, fmt):
    if range_type not in ('daily', 'monthly', 'yearly') or fmt not in ('excel', 'pdf'):
        flash('Invalid export options.', 'error')
        return redirect(url_for('toner_management'))
    user = current_user()
    if user['role'] == 'toner_req':
        return redirect(url_for('toner_req_portal'))
    ref_date = parse_date(request.args.get('date'))
    scope_all = user['role'] == 'admin' and request.args.get('scope') == 'all'
    db = get_db()
    sql = """
        SELECT toner_entry.*, user.full_name AS employee_name
        FROM toner_entry JOIN user ON toner_entry.user_id = user.id WHERE 1=1
    """
    params = []
    if not scope_all:
        sql += " AND toner_entry.user_id = ?"
        params.append(user['id'])
    sql, params = _apply_date_filter(sql, params, 'toner_entry.entry_date', range_type, ref_date)
    sql += " ORDER BY toner_entry.entry_date ASC"
    entries = db.execute(sql, params).fetchall()
    label = _range_label(range_type, ref_date)
    title = f"Toner Entries - {label}"
    headers = ["Date", "Employee", "Location", "Department", "Handover To", "Toner Type", "Qty", "Reached"]
    rows = []
    for e in entries:
        rows.append([
            _fmt_row_date(e['entry_date']), e['employee_name'] or '', e['location'] or '',
            e['department'] or '', e['handover_to'] or '', e['toner_type'] or '',
            e['quantity'] or 1, e['reached'] or '',
        ])
    filename_base = f"toner_{range_type}_{ref_date.isoformat()}"
    if fmt == 'excel':
        return _send_export(build_generic_excel(headers, rows, "Toner"), fmt, filename_base)
    return _send_export(build_generic_pdf(title, headers, rows), fmt, filename_base)


@app.route('/toner-requirement/export/<range_type>/<fmt>')
@login_required
def export_toner_requirement(range_type, fmt):
    if range_type not in ('daily', 'monthly', 'yearly') or fmt not in ('excel', 'pdf'):
        flash('Invalid export options.', 'error')
        return redirect(url_for('toner_requirement'))
    user = current_user()
    if user['role'] == 'toner_req':
        return redirect(url_for('toner_req_portal'))
    ref_date = parse_date(request.args.get('date'))
    # Requirements are shared — always export all (everyone can see all)
    db = get_db()
    sql = """
        SELECT toner_requirement.*, user.full_name AS employee_name
        FROM toner_requirement JOIN user ON toner_requirement.user_id = user.id WHERE 1=1
    """
    params = []
    sql, params = _apply_date_filter(sql, params, 'toner_requirement.entry_date', range_type, ref_date)
    sql += " ORDER BY toner_requirement.entry_date ASC"
    entries = db.execute(sql, params).fetchall()
    label = _range_label(range_type, ref_date)
    title = f"Toner Requirements - {label}"
    headers = ["Date", "Submitted By", "Needed By", "Department", "Location", "Toner Type", "Qty", "Remark", "Status"]
    rows = []
    for e in entries:
        rows.append([
            _fmt_row_date(e['entry_date']),
            e['requester_name'] or e['employee_name'] or '',
            e['needed_by'] or '', e['department'] or '', e['location'] or '',
            e['toner_type'] or '', e['quantity'] or 1, e['remark'] or '',
            e['status'] or 'open',
        ])
    filename_base = f"toner_requirement_{range_type}_{ref_date.isoformat()}"
    if fmt == 'excel':
        return _send_export(build_generic_excel(headers, rows, "Toner Req"), fmt, filename_base)
    return _send_export(build_generic_pdf(title, headers, rows), fmt, filename_base)


@app.route('/cctv/export/<range_type>/<fmt>')
@login_required
def export_cctv(range_type, fmt):
    if range_type not in ('daily', 'monthly', 'yearly') or fmt not in ('excel', 'pdf'):
        flash('Invalid export options.', 'error')
        return redirect(url_for('cctv_register'))
    user = current_user()
    if user['role'] == 'toner_req':
        return redirect(url_for('toner_req_portal'))
    ref_date = parse_date(request.args.get('date'))
    scope_all = user['role'] == 'admin' and request.args.get('scope') == 'all'
    db = get_db()
    sql = """
        SELECT cctv_entry.*, user.full_name AS employee_name
        FROM cctv_entry JOIN user ON cctv_entry.user_id = user.id WHERE 1=1
    """
    params = []
    if not scope_all:
        sql += " AND cctv_entry.user_id = ?"
        params.append(user['id'])
    sql, params = _apply_date_filter(sql, params, 'cctv_entry.entry_date', range_type, ref_date)
    sql += " ORDER BY cctv_entry.entry_date ASC"
    entries = db.execute(sql, params).fetchall()
    label = _range_label(range_type, ref_date)
    title = f"CCTV Register - {label}"
    headers = ["Date", "Employee", "User Name", "Department", "Location", "Purpose", "Remark"]
    rows = []
    for e in entries:
        rows.append([
            _fmt_row_date(e['entry_date']), e['employee_name'] or '', e['viewer_name'] or '',
            e['department'] or '', e['location'] or '', e['purpose'] or '', e['remark'] or '',
        ])
    filename_base = f"cctv_{range_type}_{ref_date.isoformat()}"
    if fmt == 'excel':
        return _send_export(build_generic_excel(headers, rows, "CCTV"), fmt, filename_base)
    return _send_export(build_generic_pdf(title, headers, rows), fmt, filename_base)


@app.route('/visitor-register/export/<range_type>/<fmt>')
@login_required
def export_visitor(range_type, fmt):
    if range_type not in ('daily', 'monthly', 'yearly') or fmt not in ('excel', 'pdf'):
        flash('Invalid export options.', 'error')
        return redirect(url_for('visitor_register'))
    user = current_user()
    if user['role'] == 'toner_req':
        return redirect(url_for('toner_req_portal'))
    ref_date = parse_date(request.args.get('date'))
    scope_all = user['role'] == 'admin' and request.args.get('scope') == 'all'
    db = get_db()
    sql = """
        SELECT visitor_entry.*, user.full_name AS employee_name
        FROM visitor_entry JOIN user ON visitor_entry.user_id = user.id WHERE 1=1
    """
    params = []
    if not scope_all:
        sql += " AND visitor_entry.user_id = ?"
        params.append(user['id'])
    sql, params = _apply_date_filter(sql, params, 'visitor_entry.entry_date', range_type, ref_date)
    sql += " ORDER BY visitor_entry.entry_date ASC"
    entries = db.execute(sql, params).fetchall()
    label = _range_label(range_type, ref_date)
    title = f"Visitor Register - {label}"
    headers = ["Date", "Employee", "Location", "Visitor Name", "Purpose", "Phone"]
    rows = []
    for e in entries:
        rows.append([
            _fmt_row_date(e['entry_date']), e['employee_name'] or '', e['location'] or '',
            e['visitor_name'] or '', e['purpose'] or '', e['phone'] or '',
        ])
    filename_base = f"visitor_{range_type}_{ref_date.isoformat()}"
    if fmt == 'excel':
        return _send_export(build_generic_excel(headers, rows, "Visitor"), fmt, filename_base)
    return _send_export(build_generic_pdf(title, headers, rows), fmt, filename_base)


init_db()


if __name__ == '__main__':
    app.run(debug=True)
